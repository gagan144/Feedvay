# Copyright (C) 2017 Feedvay (Gagandeep Singh: singh.gagan144@gmail.com) - All Rights Reserved
# Content in this document can not be copied and/or distributed without the express
# permission of Gagandeep Singh.
from openpyxl import load_workbook
import json, uuid

from storeroom.models import ImportRecord

def dump_bsp_bulk_upload(file_excel, bsp_type, org, creator_user):
    """
    Importer that dumps all BSP bulk upload records to :class:`storeroom.models.Record`.

    :param file_excel: Excel file as the source of the bulk records.
    :param bsp_type: BSP type to which records belongs to.
    :param org: Organization for which dumping is done.
    :param creator_user: User who made this import
    :return: Number of records dumped.
    """
    batch_id = str(uuid.uuid4())

    # Load workbook
    wb = load_workbook(file_excel, read_only=True)
    ws =  wb.worksheets[0]

    # get header mapping: colId-name
    map_headers = {}
    for row in ws.iter_rows():
        for col_idx, cell in enumerate(row):
            map_headers[col_idx] = cell.value
        break

    count = 0
    list_records = []
    for row_idx, row in enumerate(ws.iter_rows()):
        if row_idx == 0:
            continue

        # Obtain row json data
        row_data = {}
        for col_idx, cell in enumerate(row):
            head_name = map_headers[col_idx]
            row_data[head_name] = cell.value
        # print row_data

        # Append to bulk insert list
        list_records.append(
            ImportRecord(
                organization_id = org.id,
                batch_id = batch_id,
                context = ImportRecord.CNTX_BSP,
                filename = file_excel._name,
                identifiers = {
                    "bsp_type": bsp_type,
                },
                data = json.dumps(row_data),
                created_by = creator_user.id
            )
        )

        count += 1

    # Insert in bulk
    ImportRecord.objects.insert(list_records)

    return count